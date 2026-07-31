"""
Tabular export in the three formats the panel offers.

xlsx  — openpyxl, RTL sheet direction, frozen header, auto column widths.
csv   — UTF-8 with a BOM so Excel on Windows opens Persian text correctly.
pdf   — a self-contained, print-ready HTML document. The browser's "print to
        PDF" is used rather than a server-side PDF library: it is the only way
        to get correct Persian shaping and RTL layout without shipping and
        embedding an Arabic-script font, and it needs no new dependency.
"""
from __future__ import annotations

import csv
import html
import io
from datetime import date, datetime
from decimal import Decimal

from django.utils import timezone

Row = dict[str, object]


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "بله" if value else "خیر"
    if isinstance(value, (datetime, date)):
        return timezone.localtime(value).strftime("%Y-%m-%d %H:%M") if isinstance(
            value, datetime
        ) else value.strftime("%Y-%m-%d")
    if isinstance(value, (dict, list)):
        import json

        return json.dumps(value, ensure_ascii=False)
    return str(value)


#: Public alias — views reuse the same value formatting for JSON previews.
cell = _cell


# --------------------------------------------------------------------------
def to_csv(columns: list[tuple[str, str]], rows: list[Row]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([label for _, label in columns])
    for row in rows:
        writer.writerow([_cell(row.get(key)) for key, _ in columns])
    # BOM: Excel needs it to detect UTF-8.
    return b"\xef\xbb\xbf" + buffer.getvalue().encode("utf-8")


def to_xlsx(columns: list[tuple[str, str]], rows: list[Row], title: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or "export")[:31]
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for col, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r, row in enumerate(rows, start=2):
        for c, (key, _) in enumerate(columns, start=1):
            value = row.get(key)
            if isinstance(value, Decimal):
                value = float(value)
            elif not isinstance(value, (int, float, type(None))):
                value = _cell(value)
            ws.cell(row=r, column=c, value=value)

    for col, (key, label) in enumerate(columns, start=1):
        width = max(
            [len(label)] + [len(_cell(row.get(key))) for row in rows[:200]] or [10]
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max(width + 4, 12), 55)

    ws.freeze_panes = "A2"
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


def to_pdf_html(
    columns: list[tuple[str, str]], rows: list[Row], title: str, subtitle: str = ""
) -> bytes:
    """A print-ready A4 document that opens the print dialog on load."""
    stamp = timezone.localtime().strftime("%Y-%m-%d %H:%M")
    head = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
    body = "".join(
        "<tr>"
        + "".join(f"<td>{html.escape(_cell(row.get(key)))}</td>" for key, _ in columns)
        + "</tr>"
        for row in rows
    )
    doc = f"""<!doctype html>
<html lang="fa" dir="rtl"><head><meta charset="utf-8">
<title>{html.escape(title)}</title>
<style>
  @page {{ size: A4 landscape; margin: 12mm; }}
  * {{ box-sizing: border-box; }}
  body {{ font-family: Vazirmatn, Tahoma, sans-serif; color: #111827; margin: 0; }}
  header {{ display: flex; justify-content: space-between; align-items: baseline;
           border-bottom: 2px solid #111827; padding-bottom: 8px; margin-bottom: 14px; }}
  h1 {{ font-size: 18px; margin: 0; }}
  .sub {{ font-size: 12px; color: #6b7280; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
  th, td {{ border: 1px solid #d1d5db; padding: 5px 7px; text-align: right; }}
  th {{ background: #1f2937; color: #fff; }}
  tbody tr:nth-child(even) {{ background: #f9fafb; }}
  footer {{ margin-top: 12px; font-size: 10px; color: #9ca3af; text-align: center; }}
  @media print {{ .noprint {{ display: none; }} }}
</style></head>
<body>
  <header>
    <div><h1>{html.escape(title)}</h1>
    <div class="sub">{html.escape(subtitle)}</div></div>
    <div class="sub">تاریخ گزارش: {stamp} · {len(rows)} ردیف</div>
  </header>
  <table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>
  <footer>شرکت کاغذ حساس نمابر مهر — پنل مدیریت</footer>
  <script>window.addEventListener('load', function () {{ window.print(); }});</script>
</body></html>"""
    return doc.encode("utf-8")


CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "text/html; charset=utf-8",
}
EXTENSIONS = {"csv": "csv", "xlsx": "xlsx", "pdf": "html"}


def build(fmt: str, columns, rows, title: str, subtitle: str = "") -> tuple[bytes, str, str]:
    """-> (payload, content_type, filename). `fmt` in {csv, xlsx, pdf}."""
    fmt = (fmt or "xlsx").lower()
    if fmt not in CONTENT_TYPES:
        raise ValueError("قالب خروجی پشتیبانی نمی‌شود (csv / xlsx / pdf).")
    if fmt == "csv":
        payload = to_csv(columns, rows)
    elif fmt == "xlsx":
        payload = to_xlsx(columns, rows, title)
    else:
        payload = to_pdf_html(columns, rows, title, subtitle)
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    filename = f"{title}-{stamp}.{EXTENSIONS[fmt]}"
    return payload, CONTENT_TYPES[fmt], filename


def as_response(fmt: str, columns, rows, title: str, subtitle: str = ""):
    """Build the export and wrap it in a correctly-encoded HttpResponse."""
    from urllib.parse import quote

    from django.http import HttpResponse

    payload, content_type, filename = build(fmt, columns, rows, title, subtitle)
    response = HttpResponse(payload, content_type=content_type)
    # A "print to PDF" document must render, not download.
    disposition = "inline" if fmt == "pdf" else "attachment"
    response["Content-Disposition"] = (
        f"{disposition}; filename=export.{EXTENSIONS[fmt]}; "
        f"filename*=UTF-8''{quote(filename)}"
    )
    return response
