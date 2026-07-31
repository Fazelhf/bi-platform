"""
4 · Data management — generic CRUD over the platform's business tables, bulk
import/update/delete with validation, and the recycle bin.

Rather than 20 near-identical viewsets, one endpoint drives every table the
admin is allowed to touch: the model is named in the query string and a
ModelSerializer is built for it on the fly. The allow-list in
`services.softdelete.DELETABLE_MODELS` is the security boundary — nothing
outside it is reachable, so this cannot be pointed at auth tables.
"""
from __future__ import annotations

import base64
import binascii
import csv
import io

from django.db import transaction
from rest_framework import serializers as drf_serializers
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.viewsets import GenericViewSet, ViewSet

from apps.adminpanel.models import RecycleBin
from apps.adminpanel.permissions import AdminPanelPermission, require
from apps.adminpanel.serializers import RecycleBinSerializer
from apps.adminpanel.services import exporters, softdelete
from apps.adminpanel.services.softdelete import DELETABLE_MODELS, RecycleError
from apps.adminpanel.views.base import AdminModelViewSet
from apps.core.audit import diff, log as audit_log, snapshot
from apps.core.models import AuditLog

#: Fields never shown or written through the generic editor.
HIDDEN_FIELDS = {"created_at", "updated_at"}


def _serializer_for(model, fields: list[str] | None = None):
    """A ModelSerializer for `model`, built at request time."""
    editable = [
        f.name for f in model._meta.get_fields()
        if getattr(f, "concrete", False) and not f.auto_created
    ]
    chosen = fields or ["id"] + [f for f in editable if f != "id"]

    meta = type("Meta", (), {"model": model, "fields": chosen})
    return type(f"{model.__name__}AutoSerializer", (drf_serializers.ModelSerializer,), {"Meta": meta})


def _field_schema(model) -> list[dict]:
    """Describes each column so the UI can render the right input."""
    out = []
    for field in model._meta.get_fields():
        if not getattr(field, "concrete", False) or field.auto_created:
            continue
        if field.name in HIDDEN_FIELDS:
            continue
        kind = field.get_internal_type()
        entry = {
            "name": field.name,
            "label": str(getattr(field, "verbose_name", field.name)),
            "type": kind,
            "required": not field.blank and not field.null and not field.has_default(),
            "editable": field.editable and field.name != "id",
        }
        if field.choices:
            entry["choices"] = [
                {"value": v, "label": str(label)} for v, label in field.choices
            ]
        if field.is_relation and field.related_model:
            entry["type"] = "FK"
            entry["related"] = (
                f"{field.related_model._meta.app_label}."
                f"{field.related_model._meta.object_name}"
            )
        out.append(entry)
    return out


class DataModelViewSet(ViewSet):
    """Generic table browser/editor. `?model=<app.Model>` selects the table."""

    permission_classes = [AdminPanelPermission]
    read_permission = "data.view"
    write_permission = "data.edit"

    # -- helpers ---------------------------------------------------------
    def _model(self, request):
        label = request.query_params.get("model") or request.data.get("model")
        if not label:
            raise ValidationError({"model": "پارامتر model الزامی است."})
        try:
            return softdelete.get_model(label), label
        except RecycleError as exc:
            raise ValidationError({"model": str(exc)})

    # -- endpoints -------------------------------------------------------
    @action(detail=False, methods=["get"])
    def models(self, request):
        """Every table the panel may administer, with row counts."""
        out = []
        for label, title in DELETABLE_MODELS.items():
            model = softdelete.get_model(label)
            out.append({
                "label": label,
                "title": title,
                "app": model._meta.app_label,
                "rows": model.objects.count(),
            })
        return Response(out)

    @action(detail=False, methods=["get"])
    def schema(self, request):
        model, label = self._model(request)
        return Response({
            "model": label,
            "title": DELETABLE_MODELS[label],
            "fields": _field_schema(model),
        })

    def list(self, request):
        """Paginated rows with free-text search across char fields."""
        model, label = self._model(request)
        queryset = model.objects.all()

        search = request.query_params.get("search", "").strip()
        if search:
            from django.db.models import Q

            query = Q()
            for field in model._meta.get_fields():
                if getattr(field, "concrete", False) and field.get_internal_type() in {
                    "CharField", "TextField", "SlugField"
                }:
                    query |= Q(**{f"{field.name}__icontains": search})
            if query:
                queryset = queryset.filter(query)

        ordering = request.query_params.get("ordering")
        if ordering and ordering.lstrip("-") in {
            f.name for f in model._meta.get_fields() if getattr(f, "concrete", False)
        }:
            queryset = queryset.order_by(ordering)

        try:
            page = max(1, int(request.query_params.get("page", 1)))
            size = max(1, min(int(request.query_params.get("page_size", 50)), 500))
        except ValueError:
            page, size = 1, 50
        total = queryset.count()
        rows = queryset[(page - 1) * size: page * size]
        serializer_class = _serializer_for(model)
        return Response({
            "count": total,
            "page": page,
            "page_size": size,
            "results": serializer_class(rows, many=True).data,
        })

    def create(self, request):
        require(request.user, "data.edit")
        model, label = self._model(request)
        serializer = _serializer_for(model)(data=request.data.get("values") or {})
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        audit_log(request.user, instance, AuditLog.Action.CREATE,
                  {"created": {"before": None, "after": str(instance)[:150]}})
        return Response(serializer.data, status=201)

    def partial_update(self, request, pk=None):
        require(request.user, "data.edit")
        model, label = self._model(request)
        instance = model.objects.filter(pk=pk).first()
        if not instance:
            raise ValidationError({"detail": "رکورد یافت نشد."})
        before = snapshot(instance)
        serializer = _serializer_for(model)(
            instance, data=request.data.get("values") or {}, partial=True
        )
        serializer.is_valid(raise_exception=True)
        instance = serializer.save()
        audit_log(request.user, instance, AuditLog.Action.UPDATE,
                  diff(before, snapshot(instance)))
        return Response(serializer.data)

    def destroy(self, request, pk=None):
        """Deletes go to the recycle bin, never straight out."""
        require(request.user, "data.delete")
        model, label = self._model(request)
        instance = model.objects.filter(pk=pk).first()
        if not instance:
            raise ValidationError({"detail": "رکورد یافت نشد."})
        entry = softdelete.soft_delete(instance, request.user)
        audit_log(request.user, entry, AuditLog.Action.DELETE,
                  {"model": {"before": label, "after": None}})
        return Response({"ok": True, "recycle_id": entry.id})

    @action(detail=False, methods=["post"], url_path="bulk-delete")
    def bulk_delete(self, request):
        require(request.user, "data.delete")
        model, label = self._model(request)
        ids = request.data.get("ids") or []
        moved, errors = 0, []
        for pk in ids:
            instance = model.objects.filter(pk=pk).first()
            if not instance:
                errors.append({"id": pk, "error": "یافت نشد"})
                continue
            try:
                softdelete.soft_delete(instance, request.user)
                moved += 1
            except Exception as exc:
                errors.append({"id": pk, "error": str(exc)[:200]})
        return Response({"deleted": moved, "errors": errors})

    @action(detail=False, methods=["patch"], url_path="bulk-update")
    def bulk_update(self, request):
        require(request.user, "data.edit")
        model, label = self._model(request)
        ids = request.data.get("ids") or []
        changes = request.data.get("changes") or {}
        if not ids or not changes:
            raise ValidationError({"detail": "شناسه‌ها و تغییرات الزامی است."})
        updated, errors = 0, []
        serializer_class = _serializer_for(model)
        for instance in model.objects.filter(pk__in=ids):
            before = snapshot(instance)
            serializer = serializer_class(instance, data=changes, partial=True)
            if not serializer.is_valid():
                errors.append({"id": instance.pk, "error": serializer.errors})
                continue
            saved = serializer.save()
            audit_log(request.user, saved, AuditLog.Action.UPDATE,
                      diff(before, snapshot(saved)))
            updated += 1
        return Response({"updated": updated, "errors": errors})

    # -- import / export --------------------------------------------------
    @action(detail=False, methods=["post"], url_path="import")
    def bulk_import(self, request):
        """
        Two-phase import. `mode=validate` reports per-row errors and changes
        nothing; `mode=commit` writes only if every row is valid, inside one
        transaction — a half-imported table is never left behind.

        Input is either `rows` (list of dicts) or `file` (a data-URL of a
        .csv/.xlsx the browser picked).
        """
        require(request.user, "data.import")
        model, label = self._model(request)
        mode = request.data.get("mode", "validate")
        rows = request.data.get("rows")
        if rows is None:
            rows = self._parse_upload(request.data.get("file") or "")
        if not isinstance(rows, list) or not rows:
            raise ValidationError({"rows": "هیچ ردیفی برای ایمپورت یافت نشد."})
        if len(rows) > 5000:
            raise ValidationError({"rows": "حداکثر ۵۰۰۰ ردیف در هر ایمپورت."})

        serializer_class = _serializer_for(model)
        valid, errors = [], []
        for index, raw in enumerate(rows, start=1):
            if not isinstance(raw, dict):
                errors.append({"row": index, "errors": {"detail": "ردیف باید شیء باشد."}})
                continue
            pk = raw.get("id")
            instance = model.objects.filter(pk=pk).first() if pk else None
            payload = {k: v for k, v in raw.items() if k != "id"}
            serializer = serializer_class(instance, data=payload, partial=bool(instance))
            if serializer.is_valid():
                valid.append((index, serializer, bool(instance)))
            else:
                errors.append({"row": index, "errors": serializer.errors})

        summary = {
            "total": len(rows),
            "valid": len(valid),
            "invalid": len(errors),
            "create": sum(1 for _, _, existing in valid if not existing),
            "update": sum(1 for _, _, existing in valid if existing),
            "errors": errors[:200],
            "committed": False,
        }
        if mode != "commit":
            summary["preview"] = [s.validated_data for _, s, _ in valid[:20]]
            return Response(summary)
        if errors:
            return Response(
                {**summary, "detail": "ایمپورت انجام نشد: ابتدا خطاهای ردیف‌ها را برطرف کنید."},
                status=400,
            )
        with transaction.atomic():
            for _, serializer, existing in valid:
                instance = serializer.save()
                audit_log(request.user, instance,
                          AuditLog.Action.UPDATE if existing else AuditLog.Action.CREATE,
                          {"import": {"before": None, "after": "bulk"}})
        summary["committed"] = True
        return Response(summary)

    @staticmethod
    def _parse_upload(data_url: str) -> list[dict]:
        """Decode a data-URL holding a .csv or .xlsx into a list of row dicts."""
        if not data_url.startswith("data:"):
            raise ValidationError({"file": "فایل باید به صورت data-URL ارسال شود."})
        header, _, payload = data_url.partition(",")
        try:
            blob = base64.b64decode(payload)
        except (binascii.Error, ValueError):
            raise ValidationError({"file": "محتوای فایل قابل خواندن نیست."})

        if "sheet" in header or "excel" in header or "xlsx" in header:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
            ws = wb.active
            iterator = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h is not None else "" for h in next(iterator, [])]
            return [
                {h: v for h, v in zip(headers, values) if h}
                for values in iterator
                if any(v is not None for v in values)
            ]

        text = blob.decode("utf-8-sig", errors="replace")
        return [dict(row) for row in csv.DictReader(io.StringIO(text))]

    @action(detail=False, methods=["get"])
    def export(self, request):
        """?model=<app.Model>&fmt=xlsx|csv|pdf (`fmt`, since DRF owns `format`)."""
        require(request.user, "data.export")
        model, label = self._model(request)
        fmt = request.query_params.get("fmt", "xlsx")
        rows = _serializer_for(model)(model.objects.all()[:10000], many=True).data
        columns = [
            (f["name"], f["label"]) for f in _field_schema(model)
        ]
        try:
            return exporters.as_response(
                fmt, columns, list(rows), DELETABLE_MODELS[label]
            )
        except ValueError as exc:
            raise ValidationError({"format": str(exc)})

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """An empty xlsx with the right headers — the starting point for imports."""
        require(request.user, "data.import")
        model, label = self._model(request)
        columns = [
            (f["name"], f["label"]) for f in _field_schema(model) if f["editable"]
        ]
        return exporters.as_response(
            "xlsx", columns, [], f"الگوی {DELETABLE_MODELS[label]}"
        )


# ==========================================================================
# Recycle bin
# ==========================================================================
class RecycleBinViewSet(AdminModelViewSet):
    queryset = RecycleBin.objects.select_related("deleted_by", "restored_by")
    serializer_class = RecycleBinSerializer
    read_permission = "data.view"
    write_permission = "data.restore"
    filterset_fields = ["model_label", "deleted_by"]
    search_fields = ["object_repr", "model_label", "note"]
    ordering_fields = ["deleted_at", "restored_at"]
    http_method_names = ["get", "post", "delete", "head", "options"]
    export_title = "سطل بازیافت"
    export_columns = [
        ("model_label_fa", "نوع"), ("object_repr", "رکورد"),
        ("deleted_by_name", "حذف توسط"), ("deleted_at", "زمان حذف"),
        ("restored_at", "زمان بازیابی"),
    ]

    def get_queryset(self):
        qs = super().get_queryset()
        state = self.request.query_params.get("state")
        if state == "pending":
            qs = qs.filter(restored_at__isnull=True)
        elif state == "restored":
            qs = qs.filter(restored_at__isnull=False)
        return qs

    @action(detail=True, methods=["post"])
    def restore(self, request, pk=None):
        require(request.user, "data.restore")
        entry = self.get_object()
        try:
            obj = softdelete.restore(entry, request.user)
        except RecycleError as exc:
            raise ValidationError({"detail": str(exc)})
        audit_log(request.user, obj, AuditLog.Action.CREATE,
                  {"restored_from_bin": {"before": None, "after": str(entry.id)}})
        return Response(self.get_serializer(entry).data)

    @action(detail=False, methods=["post"], url_path="purge")
    def purge(self, request):
        """Permanently forget binned rows. Irreversible — hence its own action."""
        require(request.user, "data.delete")
        ids = request.data.get("ids")
        queryset = self.get_queryset().filter(restored_at__isnull=True)
        if ids:
            queryset = queryset.filter(pk__in=ids)
        elif not request.data.get("all"):
            raise ValidationError({"detail": "ids یا all=true لازم است."})
        count = queryset.count()
        queryset.delete()
        return Response({"purged": count})


class DataOverviewView(APIView):
    """Row counts per table + bin summary, for the data page header."""

    permission_classes = [AdminPanelPermission]
    read_permission = "data.view"

    def get(self, request):
        tables = []
        for label, title in DELETABLE_MODELS.items():
            model = softdelete.get_model(label)
            tables.append({
                "label": label, "title": title,
                "app": model._meta.app_label, "rows": model.objects.count(),
            })
        return Response({
            "tables": sorted(tables, key=lambda t: t["rows"], reverse=True),
            "recycle_pending": RecycleBin.objects.filter(restored_at__isnull=True).count(),
            "recycle_restored": RecycleBin.objects.filter(restored_at__isnull=False).count(),
        })
