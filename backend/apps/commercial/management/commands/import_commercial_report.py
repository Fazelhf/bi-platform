"""
Load the department's real Commercial Report workbook.

    python manage.py import_commercial_report "Commercial Report1.xlsx"
    python manage.py import_commercial_report FILE --replace   # wipe first
    python manage.py import_commercial_report FILE --dry-run

A command rather than a data migration: this is one company's operational
file, not reference data, and `deploy.sh` runs migrations automatically.

**What the workbook is, and why it is read this way.** It keeps five tables
for what is really one thing — منتظر مجوز, ثبت‌شده, آماده برای تخصیص,
اظهارنشده, ترخیص‌شده — and a file is cut and pasted from one to the next as
it progresses. That is why several of them disagree: WL2025-19 appears in
three, with a different وزن and a different نوع محموله in each. So the
importer reads them in **pipeline order** and lets later, more advanced
tables correct earlier ones, keeping whichever value is most specific rather
than whichever it saw last.

Two date formats live side by side in the same columns — «1404/07/16» is
Jalali text, `2025-12-11 00:00:00` is a real datetime, and «7/20/2026» is a
US-format string someone typed. All three appear in تاریخ انقضا پروفرما
alone. `_date` handles all of them and returns None rather than guessing.
"""
from __future__ import annotations

import datetime as dt
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.text import slugify

from apps.commercial.models import (
    Bank,
    Currency,
    ForeignOrder,
    Shipment,
    Supplier,
)
from apps.core import jalali

JALALI_RE = re.compile(r"^\s*(1[34]\d{2})[/-](\d{1,2})[/-](\d{1,2})\s*$")
US_RE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")
#: «در تاریخ 1404/03/06 بار آزاد شد» — the release date hidden in a sentence.
INLINE_JALALI = re.compile(r"(1[34]\d{2})/(\d{1,2})/(\d{1,2})")


def _date(value) -> dt.date | None:
    """A date from whatever the cell happens to hold, or None."""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text or text == "/":
        return None

    m = JALALI_RE.match(text)
    if m:
        try:
            return jalali.to_gregorian(int(m[1]), int(m[2]), int(m[3]))
        except (ValueError, IndexError):
            return None
    m = US_RE.match(text)
    if m:
        try:
            return dt.date(int(m[3]), int(m[1]), int(m[2]))
        except ValueError:
            return None
    return None


def _num(value) -> Decimal:
    if value is None or value == "":
        return Decimal(0)
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = re.sub(r"[^\d.\-]", "", str(value))
    if not text or text in {"-", ".", "-."}:
        return Decimal(0)
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal(0)


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _pi(value) -> str:
    """
    The PI number, stripped of the notes people write beside it.

    «WL2025-19- ثبت آماری» and «ثبت آماری WL2025-42» are the same kind of
    thing as «WL2025-19»: a PI with a comment attached. Matching on the raw
    string would create three files for one shipment.
    """
    text = _text(value)
    if not text:
        return ""
    text = text.replace("ثبت آماری", "").replace("ث آماری", "")
    text = text.strip(" -–—‌")
    return text


class Command(BaseCommand):
    help = "بارگذاری فایل Commercial Report واقعی بازرگانی"

    def add_arguments(self, parser):
        parser.add_argument("path", help="مسیر فایل .xlsx")
        parser.add_argument(
            "--replace", action="store_true",
            help="حذف پرونده‌ها و محموله‌های موجود پیش از بارگذاری",
        )
        parser.add_argument(
            "--dry-run", action="store_true",
            help="فقط گزارش کن، چیزی ننویس",
        )

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:
            raise CommandError("openpyxl نصب نیست: pip install openpyxl")

        try:
            wb = openpyxl.load_workbook(opts["path"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"فایل پیدا نشد: {opts['path']}")

        self.banks: dict[str, Bank] = {
            b.name_fa: b for b in Bank.objects.all()
        }
        self.seller = None
        self.orders: dict[str, ForeignOrder] = {}
        self.counts = {"orders": 0, "shipments": 0, "skipped": 0}

        with transaction.atomic():
            if opts["replace"]:
                removed = Shipment.objects.all().delete()[0]
                removed += ForeignOrder.objects.all().delete()[0]
                self.stdout.write(f"حذف شد: {removed} ردیف")

            self.seller = self._seller()
            main = wb["Commercial Report"]

            # Pipeline order: earliest stage first, so a later table showing
            # the same PI further along overwrites the status rather than the
            # other way round.
            self._permits(main)
            self._registered(main)
            self._queue(main)
            self._undeclared(main)
            self._clearance(main)
            self._shipments(main)

            if opts["dry_run"]:
                self.stdout.write(self.style.WARNING("dry-run — برگشت داده شد"))
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(
            f"پرونده: {self.counts['orders']} · "
            f"محموله: {self.counts['shipments']} · "
            f"نادیده: {self.counts['skipped']}"
        ))

    # -- helpers ---------------------------------------------------------
    def _seller(self) -> Supplier:
        supplier, _ = Supplier.objects.get_or_create(
            code="oriental-paper",
            defaults={
                "name_fa": "اورینتال پیپر", "name_en": "Oriental Paper",
                "origin": Supplier.Origin.FOREIGN, "country": "چین",
                "activity": "تولید کاغذ حرارتی و تحریر",
            },
        )
        return supplier

    def _bank(self, name: str) -> Bank | None:
        name = _text(name)
        if not name:
            return None
        if name in self.banks:
            return self.banks[name]
        bank = Bank.objects.create(
            code=slugify(name, allow_unicode=True)[:50] or f"bank-{len(self.banks)}",
            name_fa=name, sort_order=50 + len(self.banks),
        )
        self.banks[name] = bank
        return bank

    def _order(self, pi: str, **fields) -> ForeignOrder | None:
        """
        Create or advance a file.

        Blank incoming values never overwrite something already known: the
        earlier tables carry less detail, and a later one leaving a cell empty
        means "unchanged", not "cleared".
        """
        pi = _pi(pi)
        if not pi:
            self.counts["skipped"] += 1
            return None

        order = self.orders.get(pi)
        if order is None:
            order = ForeignOrder.objects.filter(pi_no=pi).first()
        if order is None:
            order = ForeignOrder(pi_no=pi, supplier=self.seller, country="چین")
            self.counts["orders"] += 1

        for key, value in fields.items():
            if value in (None, "", Decimal(0)) and getattr(order, key, None):
                continue
            if value in (None, ""):
                continue
            setattr(order, key, value)

        order.save()
        self.orders[pi] = order
        return order

    @staticmethod
    def _rows(ws, header_row: int, stop_on_blank: bool = True):
        """Rows under a header, stopping at the first blank or total line."""
        for row in ws.iter_rows(min_row=header_row + 1):
            first = _text(row[0].value)
            if not first:
                if stop_on_blank:
                    return
                continue
            if first.upper().startswith("TOTAL"):
                return
            yield row

    @staticmethod
    def _find(ws, needle: str) -> int | None:
        for row in ws.iter_rows(min_col=1, max_col=1):
            if needle in _text(row[0].value):
                return row[0].row
        return None

    # -- the five stage tables -------------------------------------------
    def _permits(self, ws):
        """«ثبت سفارش های انجام شده - منتظر مجوز صنعت و معدن»."""
        head = self._find(ws, "منتظر مجوز صنعت و معدن")
        if not head:
            return
        for row in self._rows(ws, head + 1):
            self._order(
                row[0].value,
                registration_no=_text(row[1].value),
                goods_desc=_text(row[2].value),
                weight_ton=_num(row[3].value),
                amount=_num(row[4].value),
                bank=self._bank(row[5].value),
                currency=Currency.USD,
                status=ForeignOrder.Status.AWAITING_PERMIT,
                last_status_note=_text(row[6].value),
            )

    def _registered(self, ws):
        """«ثبت سفارش های انجام شده که در نوبت تخصیص قرار نگرفته اند»."""
        head = self._find(ws, "در نوبت تخصیص قرار نگرفته")
        if not head:
            return
        for row in self._rows(ws, head + 1):
            self._order(
                row[0].value,
                registration_no=_text(row[1].value),
                goods_desc=_text(row[2].value),
                weight_ton=_num(row[3].value),
                amount=_num(row[4].value),
                bank=self._bank(row[5].value),
                registered_on=_date(row[6].value),
                valid_until=_date(row[7].value),
                proforma_expires_on=_date(row[8].value),
                insurance=_text(row[9].value) or ForeignOrder.Readiness.PENDING,
                inspection=_text(row[10].value) or ForeignOrder.Readiness.PENDING,
                status=ForeignOrder.Status.REGISTERED,
            )

    def _queue(self, ws):
        """«ثبت سفارش های آماده برای تخصیص» — the allocation queue itself."""
        head = self._find(ws, "ثبت سفارش های آماده برای تخصیص")
        if not head:
            return
        for row in self._rows(ws, head + 1):
            expected = int(_num(row[11].value)) or 100
            self._order(
                row[0].value,
                registration_no=_text(row[1].value),
                goods_desc=_text(row[2].value),
                weight_ton=_num(row[3].value),
                amount=_num(row[4].value),
                bank=self._bank(row[5].value),
                registered_on=_date(row[6].value),
                valid_until=_date(row[7].value),
                proforma_expires_on=_date(row[8].value),
                # «تاریخ تایید در سامانه» is the day it entered the queue —
                # the workbook's «تعداد روز انتظار» counts from here.
                queued_on=_date(row[9].value),
                expected_queue_days=expected,
                insurance=_text(row[14].value) or ForeignOrder.Readiness.PENDING,
                inspection=_text(row[15].value) or ForeignOrder.Readiness.PENDING,
                status=ForeignOrder.Status.QUEUED,
                last_status_note=_text(row[13].value),
            )

    def _undeclared(self, ws):
        """«سفارشهایی که اظهار گمرکی نشده اند» — landed, not yet declared."""
        head = self._find(ws, "اظهار گمرکی نشده")
        if not head:
            return
        for row in self._rows(ws, head + 1):
            self._order(
                row[0].value,
                registration_no=_text(row[1].value),
                goods_desc=_text(row[2].value),
                weight_ton=_num(row[3].value),
                amount=_num(row[4].value),
                bank=self._bank(row[5].value),
                registered_on=_date(row[6].value),
                valid_until=_date(row[7].value),
                proforma_expires_on=_date(row[8].value),
                arrived_on=_date(row[9].value),
                factory_entry_on=_date(row[10].value),
                production_cert_on=_date(row[11].value),
                customs_declared_on=_date(row[12].value),
                status=ForeignOrder.Status.CUSTOMS,
                last_status_note=_text(row[14].value),
            )

    def _clearance(self, ws):
        """«بارهای موجود در گمرک و در راه و ترخیص شده»."""
        head = self._find(ws, "بارهای موجود در گمرک و در راه")
        if not head:
            return
        for row in self._rows(ws, head + 1):
            note = _text(row[6].value)
            cleared = _date(row[7].value)
            # «ترخیص شده» / «90% ترخیص شده» / «در بندر-اظهار 1404» — the
            # sentence is the only place the outcome is recorded.
            if cleared or "ترخیص شده" in note:
                status = ForeignOrder.Status.CLEARED
            elif "در بندر" in note:
                status = ForeignOrder.Status.CUSTOMS
            else:
                status = None
            self._order(
                row[0].value,
                weight_ton=_num(row[1].value),
                goods_desc=_text(row[2].value),
                arrived_on=_date(row[3].value),
                cleared_on=cleared,
                status=status,
                last_status_note=note,
            )

    def _shipments(self, ws):
        """The main table: one row per بارنامه, with money and dates."""
        head = self._find(ws, "شماره  PI")
        if not head:
            return
        # This table has a blank continuation row in the middle — the second
        # interest instalment for FP25080709 sits on its own line with no PI.
        # Stopping at the first blank would drop two thirds of the shipments,
        # so it runs to the TOTAL row instead.
        for row in self._rows(ws, head, stop_on_blank=False):
            pi_raw = _pi(row[0].value)
            if not pi_raw:
                continue
            # «FP25021913TK-01RM-2» is lot 2 of file «FP25021913TK-01RM».
            base, _, lot = pi_raw.rpartition("-")
            if base and lot.isdigit() and len(lot) <= 2:
                file_pi, lot_no = base, lot
            else:
                file_pi, lot_no = pi_raw, ""

            order = self._order(
                file_pi,
                goods_desc=_text(row[3].value),
                currency=Currency.USD,
            )
            if order is None:
                continue

            arrived = _date(row[13].value)
            cleared = _date(row[17].value)
            note = _text(row[15].value)
            released = _date(row[12].value)
            if not released:
                m = INLINE_JALALI.search(note)
                if m:
                    released = _date(f"{m[1]}/{m[2]}/{m[3]}")

            clearance = _text(row[16].value)
            if cleared or "ترخیص شد" in clearance:
                status = Shipment.Status.CLEARED
            elif arrived:
                status = Shipment.Status.CUSTOMS
            else:
                status = Shipment.Status.AT_SEA

            bl = _text(row[1].value)
            shipment, created = Shipment.objects.update_or_create(
                order=order,
                lot_no=lot_no,
                bl_no=bl,
                defaults={
                    "weight_ton": _num(row[2].value),
                    "goods_desc": _text(row[3].value),
                    "value_amount": _num(row[4].value),
                    "paid_amount": _num(row[5].value),
                    "due_on": _date(row[7].value),
                    "interest_amount": _num(row[9].value),
                    "etd": _date(row[10].value),
                    "eta": _date(row[11].value),
                    "released_on": released,
                    "arrived_on": arrived,
                    "cleared_on": cleared,
                    "status": status,
                    "note": note,
                },
            )
            if created:
                self.counts["shipments"] += 1
