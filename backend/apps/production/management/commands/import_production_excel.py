"""
Import the تولید (Production) KPI workbook into the warehouse.

Reads raw *input* cells only (ورودی، منابع، درامد، مجموع) rather than cached
formula results, so the import does not depend on what Excel last calculated.

Usage:
    python manage.py import_production_excel \
        --file "KPI اردیبهشت تولید1405.xlsx" --year 1405 --month 2 [--approve]

DEFECT FOUND IN SOURCE: the per-machine sheets (دستگاه ۱–۵) map downtime
reasons to the wrong ورودی rows — 'خرابی' reads row 8 (which is تغییر سایز)
and 'تغییر سایز' reads row 7 (which is خرابی). The ورودی labels are
corroborated by the sheet's own notes (H3 "شامل ۰/۵ شیفت تغییر سایز",
H9 "عدم وجود سفارش"), so ورودی is treated as authoritative here.
"""
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.core.models import DimPeriod
from apps.production.models import (
    DimCostCategory,
    DimMachine,
    DimProduct,
    FactMaterialBalance,
    FactPrintColor,
    FactProduction,
    FactProductionCost,
    FactProductionRevenue,
    ProductionBenchmark,
)
from apps.production.services.kpi import compute_period_kpis
from apps.sales.models import ApprovalStatus

# ورودی: cutting lines occupy columns C..G (3..7), machine code by column.
CUTTING_COLS = {3: "cut-1", 4: "cut-2", 5: "cut-3", 6: "cut-4", 7: "cut-5"}
CUT_ROWS = {
    "active_shifts": 3,
    "output_units": 4,
    "waste_pct": 5,
    "repair_count": 6,
    "downtime_breakdown_shifts": 7,   # ورودی label: خرابی
    "downtime_sizechange_shifts": 8,  # ورودی label: تغییر سایز
    "downtime_nowork_shifts": 9,      # ورودی label: عدم کار
}

# ورودی print block (totals in column G = 7).
PRINT_ROWS = {
    "active_shifts": 12,
    "output_units": 13,
    "waste_pct": 14,
    "downtime_breakdown_shifts": 15,
    "downtime_sizechange_shifts": 16,
    "downtime_nowork_shifts": 17,
}
PRINT_COLOR_COLS = {3: 1, 4: 2, 5: 3, 6: 4}  # C..F -> 1..4 colours, row 13

# Costs: (category code, sheet, column, row). منابع holds manual overrides
# that the workbook's own total (منابع!D4) actually uses.
COST_SOURCES = [
    ("cost-production", "منابع", 3, 4),
    ("cost-rent", "منابع", 3, 5),
    ("cost-maintenance", "ورودی", 11, 5),
    ("cost-utilities", "ورودی", 11, 6),
    ("cost-salary", "منابع", 3, 8),
    ("cost-transport", "ورودی", 11, 8),
    ("cost-other", "ورودی", 11, 9),
]

# Revenue: (product code, ورودی quantity cell row in col J=10, درامد rate row).
REVENUE_SOURCES = [
    ("prod-57", 12, 4),
    ("prod-79", 13, 5),
    ("prod-receipt", 14, 6),
    ("prod-print", 15, 7),
]


def _num(value) -> Decimal:
    if value is None or value in ("", "-"):
        return Decimal(0)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal(0)


class Command(BaseCommand):
    help = "Import the Production KPI workbook."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--year", type=int, required=True)
        parser.add_argument("--month", type=int, required=True)
        parser.add_argument("--approve", action="store_true")

    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["file"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {opts['file']}")

        for sheet in ("ورودی", "منابع", "درامد", "مجموع"):
            if sheet not in wb.sheetnames:
                raise CommandError(f"Missing sheet '{sheet}'. Found: {wb.sheetnames}")

        wsin, wsres, wsrev, wssum = (
            wb["ورودی"], wb["منابع"], wb["درامد"], wb["مجموع"]
        )
        period, _ = DimPeriod.objects.get_or_create(
            jalali_year=opts["year"], jalali_month=opts["month"]
        )
        status = ApprovalStatus.APPROVED if opts["approve"] else ApprovalStatus.DRAFT

        # ---- Benchmarks (مجموع H3..H6 + ورودی K2) ----
        bench, _ = ProductionBenchmark.objects.get_or_create(period=period)
        bench.ideal_output_per_shift = int(_num(wssum.cell(3, 8).value) or 16000)
        bench.monthly_shift_capacity = int(_num(wssum.cell(4, 8).value) or 120)
        bench.hours_per_shift = int(_num(wssum.cell(5, 8).value) or 8)
        bench.full_system_staff = int(_num(wssum.cell(6, 8).value) or 33)
        bench.total_headcount = int(_num(wsin.cell(2, 11).value) or 675)
        bench.save()

        # ---- Cutting lines ----
        machines = {m.code: m for m in DimMachine.objects.all()}
        n_machines = 0
        for col, code in CUTTING_COLS.items():
            machine = machines.get(code)
            if machine is None:
                continue
            values = {
                field: _num(wsin.cell(row, col).value)
                for field, row in CUT_ROWS.items()
            }
            FactProduction.objects.update_or_create(
                period=period, machine=machine,
                defaults={**values, "status": status},
            )
            n_machines += 1

        # ---- Print unit (totals live in column G) ----
        if "print" in machines:
            values = {
                field: _num(wsin.cell(row, 7).value)
                for field, row in PRINT_ROWS.items()
            }
            FactProduction.objects.update_or_create(
                period=period, machine=machines["print"],
                defaults={**values, "repair_count": 0, "status": status},
            )
            n_machines += 1

        # ---- Print colour breakdown (row 13, cols C..F) ----
        n_colors = 0
        for col, colors in PRINT_COLOR_COLS.items():
            area = _num(wsin.cell(13, col).value)
            if not area:
                continue
            FactPrintColor.objects.update_or_create(
                period=period, color_count=colors, defaults={"area_sqm": area}
            )
            n_colors += 1

        # ---- Costs ----
        sheets = {"ورودی": wsin, "منابع": wsres}
        categories = {c.code: c for c in DimCostCategory.objects.all()}
        n_costs = 0
        for code, sheet_name, col, row in COST_SOURCES:
            category = categories.get(code)
            if category is None:
                continue
            FactProductionCost.objects.update_or_create(
                period=period, category=category,
                defaults={"amount_rial": _num(sheets[sheet_name].cell(row, col).value)},
            )
            n_costs += 1

        # ---- Revenue (quantities from ورودی col J, rates from درامد col D) ----
        products = {p.code: p for p in DimProduct.objects.all()}
        n_rev = 0
        for code, qty_row, rate_row in REVENUE_SOURCES:
            product = products.get(code)
            if product is None:
                continue
            rate = _num(wsrev.cell(rate_row, 4).value) or product.piece_rate_rial
            FactProductionRevenue.objects.update_or_create(
                period=period, product=product,
                defaults={
                    "quantity": _num(wsin.cell(qty_row, 10).value),
                    "piece_rate_rial": rate,
                },
            )
            n_rev += 1

        # ---- Material balance (مجموع rows 12/13 cutting, 15/16 print) ----
        for stream, in_row, out_row in (
            (FactMaterialBalance.Stream.CUTTING, 12, 13),
            (FactMaterialBalance.Stream.PRINT, 15, 16),
        ):
            FactMaterialBalance.objects.update_or_create(
                period=period, stream=stream,
                defaults={
                    "input_weight": _num(wssum.cell(in_row, 3).value),
                    "output_weight": _num(wssum.cell(out_row, 3).value),
                },
            )

        self.stdout.write(self.style.SUCCESS(
            f"Imported {n_machines} machines, {n_colors} print-colour rows, "
            f"{n_costs} cost lines, {n_rev} revenue lines for {period.label} ({status})."
        ))

        if opts["approve"]:
            written = compute_period_kpis(period)
            self.stdout.write(self.style.SUCCESS(f"Computed {written} production KPI rows."))
