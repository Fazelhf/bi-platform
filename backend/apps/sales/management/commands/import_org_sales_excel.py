"""
Import the سازمانی (Organizational Sales) workbook — the second sales channel.

This is the key-account / بانکی channel: two named salespeople with large
per-person volumes, provincial sales, and bank-collection figures. Loading it
fills the gap that made the بانکی team read zero (that data lives here, not in
the همکار workbook).

Usage:
    python manage.py import_org_sales_excel \
        --file "سازمانیKPI ورودی اردیبهشت 1405.xlsx" --year 1405 --month 2 [--approve]
"""
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.core.models import DimPeriod
from apps.sales.models import (
    ApprovalStatus,
    DimBank,
    DimEmployee,
    DimProvince,
    DimTeam,
    FactCollection,
    FactSalesMonthly,
    FactSalesProvince,
    SalesChannel,
)
from apps.sales.services.kpi import compute_period_kpis

NAME_ROW = 2
FIRST_DATA_COL = 3  # column C
MEASURE_ROWS = {
    "revenue_rial": 3,
    "invoice_count": 4,
    "active_customers": 5,
    "new_customers": 6,
    "profit_rial": 7,
    "cost_rial": 8,
    "target_rial": 9,
    "calls": 10,
}
# Org salespeople belong to the بانکی channel/team.
ORG_TEAM_CODE = "banking"

# Province block spans rows 14..48 (Tehran sits separately at 48).
PROVINCE_ROWS = range(15, 49)
PROVINCE_NAME_COL, PROVINCE_SALES_COL, PROVINCE_TARGET_COL = 2, 3, 4

# Bank block: names in column G (7), collected amounts in column H (8).
BANK_NAME_COL, BANK_AMOUNT_COL = 7, 8
BANK_ROWS = range(15, 44)


def _num(value) -> Decimal:
    if value is None or value in ("", "-"):
        return Decimal(0)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal(0)


def _int(value) -> int:
    return int(_num(value))


class Command(BaseCommand):
    help = "Import the Organizational Sales workbook (channel=organizational)."

    def add_arguments(self, parser):
        parser.add_argument("--file", required=True)
        parser.add_argument("--sheet", default="Sheet1")
        parser.add_argument("--year", type=int, required=True)
        parser.add_argument("--month", type=int, required=True)
        parser.add_argument("--approve", action="store_true")

    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["file"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {opts['file']}")
        ws = wb[opts["sheet"]] if opts["sheet"] in wb.sheetnames else wb.worksheets[0]

        period, _ = DimPeriod.objects.get_or_create(
            jalali_year=opts["year"], jalali_month=opts["month"]
        )
        status = ApprovalStatus.APPROVED if opts["approve"] else ApprovalStatus.DRAFT
        banking = DimTeam.objects.filter(code=ORG_TEAM_CODE).first()

        # ---- Salespeople (skip empty فروشنده placeholders) ----
        n_facts = 0
        col = FIRST_DATA_COL
        while True:
            name = ws.cell(row=NAME_ROW, column=col).value
            if not name or str(name).strip() in ("جمع", ""):
                break
            name = str(name).strip()
            col += 1
            # Skip growth placeholders like "فروشنده1" with no data.
            revenue = _num(ws.cell(row=MEASURE_ROWS["revenue_rial"], column=col - 1).value)
            if name.startswith("فروشنده") and not revenue:
                continue

            employee, _ = DimEmployee.objects.get_or_create(
                full_name_fa=name,
                defaults={"code": f"org-{col}", "team": banking},
            )
            values = {
                f: (_num if "rial" in f else _int)(ws.cell(row=r, column=col - 1).value)
                for f, r in MEASURE_ROWS.items()
            }
            FactSalesMonthly.objects.update_or_create(
                period=period, employee=employee,
                channel=SalesChannel.ORGANIZATIONAL,
                defaults={**values, "status": status},
            )
            n_facts += 1

        # ---- Provinces (organizational channel overwrites its own rows) ----
        prov_index = {p.name_fa: p for p in DimProvince.objects.all()}
        n_prov = 0
        for row in PROVINCE_ROWS:
            name = ws.cell(row=row, column=PROVINCE_NAME_COL).value
            if not name or str(name).strip() in ("استانها", ""):
                continue
            name = str(name).strip()
            sales = _num(ws.cell(row=row, column=PROVINCE_SALES_COL).value)
            target = _num(ws.cell(row=row, column=PROVINCE_TARGET_COL).value)
            province = prov_index.get(name) or DimProvince.objects.create(
                code=f"prov-org-{row}", name_fa=name
            )
            prov_index[name] = province
            # Channel-scoped: the organizational channel owns its own province
            # row; dashboards sum team + organizational for the combined map.
            FactSalesProvince.objects.update_or_create(
                period=period, province=province,
                channel=SalesChannel.ORGANIZATIONAL,
                defaults={"sales_rial": sales, "target_rial": target},
            )
            n_prov += 1

        # ---- Bank collections (only rows carrying an amount) ----
        n_banks = 0
        for row in BANK_ROWS:
            bank_name = ws.cell(row=row, column=BANK_NAME_COL).value
            amount = _num(ws.cell(row=row, column=BANK_AMOUNT_COL).value)
            if not bank_name or not amount:
                continue
            bank_name = str(bank_name).strip()
            bank, _ = DimBank.objects.get_or_create(
                name_fa=bank_name,
                defaults={"code": f"bank-{row}", "kind": DimBank.Kind.BANK},
            )
            FactCollection.objects.update_or_create(
                period=period, bank=bank, defaults={"amount_rial": amount}
            )
            n_banks += 1

        self.stdout.write(self.style.SUCCESS(
            f"Imported {n_facts} org salespeople, {n_prov} provinces, "
            f"{n_banks} bank collections for {period.label} ({status})."
        ))

        if opts["approve"]:
            written = compute_period_kpis(period)
            self.stdout.write(self.style.SUCCESS(f"Recomputed {written} sales KPI rows."))
