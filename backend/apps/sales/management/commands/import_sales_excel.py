"""
Import the company's real KPI workbook into the warehouse.

This reads the "همکار / Employee KPI" workbook (the finest-grain source:
one column per salesperson) and loads FactSalesMonthly + FactSalesProvince,
creating employee/province dimensions on the fly and mapping each
salesperson to the team they belong to in Employee!Sheet3.

Usage:
    python manage.py import_sales_excel \
        --employee-file "KPI همکار اردیبهشت 1405 (1).xlsx" \
        --year 1405 --month 2 [--approve]

Rows are imported as DRAFT by default (so the approval workflow can be
demoed); pass --approve to import them already approved and compute KPIs.
"""
from decimal import Decimal, InvalidOperation

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from apps.core.models import DimPeriod
from apps.sales.models import (
    ApprovalStatus,
    DimEmployee,
    DimProvince,
    DimTeam,
    FactSalesMonthly,
    FactSalesProvince,
    SalesChannel,
)
from apps.sales.services.kpi import compute_period_kpis

# Salesperson -> team (derived from the aggregation columns in Sheet3).
NAME_TO_TEAM = {
    "هانیه منزه": "banking",
    "مهسا احمدی": "banking",
    "هستی خانی": "west",
    "مهدیس مومنی": "west",
    "افسانه چوبینی": "east",
    "مهسا قنبری": "east",
    "صبا موسوی": "tehran",
    "پارسا مروتی": "tehran",
    "شیما نظام ابادی": "b2b",
}

# Row layout of the input sheet (1-indexed), measure -> row.
MEASURE_ROWS = {
    "revenue_rial": 3,
    "invoice_count": 4,
    "active_customers": 5,
    "new_customers": 6,
    "profit_rial": 7,
    "cost_rial": 8,
    "target_rial": 9,
    "calls": 10,
}
NAME_ROW = 2
FIRST_DATA_COL = 3  # column C
PROVINCE_FIRST_ROW = 15
PROVINCE_LAST_ROW = 45
PROVINCE_NAME_COL = 2  # B
PROVINCE_SALES_COL = 3  # C
PROVINCE_TARGET_COL = 4  # D


def _num(value) -> Decimal:
    if value is None or value == "" or value == "-":
        return Decimal(0)
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return Decimal(0)


def _int(value) -> int:
    return int(_num(value))


class Command(BaseCommand):
    help = "Import the Employee KPI workbook into the sales warehouse."

    def add_arguments(self, parser):
        parser.add_argument("--employee-file", required=True)
        parser.add_argument("--sheet", default="ورودی")
        parser.add_argument("--year", type=int, required=True)
        parser.add_argument("--month", type=int, required=True)
        parser.add_argument("--approve", action="store_true",
                            help="Import rows as approved and compute KPIs immediately.")

    def handle(self, *args, **opts):
        try:
            wb = openpyxl.load_workbook(opts["employee_file"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {opts['employee_file']}")
        if opts["sheet"] not in wb.sheetnames:
            raise CommandError(f"Sheet '{opts['sheet']}' not in {wb.sheetnames}")
        ws = wb[opts["sheet"]]

        period, _ = DimPeriod.objects.get_or_create(
            jalali_year=opts["year"], jalali_month=opts["month"]
        )
        teams = {t.code: t for t in DimTeam.objects.all()}
        status = ApprovalStatus.APPROVED if opts["approve"] else ApprovalStatus.DRAFT

        # ---- Salespeople (columns) ----
        emp_cols = []  # (col_index, DimEmployee)
        col = FIRST_DATA_COL
        idx = 0
        while True:
            name = ws.cell(row=NAME_ROW, column=col).value
            if not name or str(name).strip() in ("جمع", ""):
                break
            name = str(name).strip()
            idx += 1
            employee, _ = DimEmployee.objects.get_or_create(
                full_name_fa=name,
                defaults={
                    "code": f"emp-{idx}",
                    "team": teams.get(NAME_TO_TEAM.get(name)),
                },
            )
            # Backfill team if it was created without one.
            if employee.team_id is None and NAME_TO_TEAM.get(name):
                employee.team = teams.get(NAME_TO_TEAM[name])
                employee.save(update_fields=["team"])
            emp_cols.append((col, employee))
            col += 1

        # ---- Facts: one row per salesperson ----
        n_facts = 0
        for col, employee in emp_cols:
            values = {
                field: (_num if "rial" in field else _int)(
                    ws.cell(row=row, column=col).value
                )
                for field, row in MEASURE_ROWS.items()
            }
            FactSalesMonthly.objects.update_or_create(
                period=period,
                employee=employee,
                channel=SalesChannel.TEAM,
                defaults={**values, "status": status},
            )
            n_facts += 1

        # ---- Province block ----
        n_prov = 0
        prov_index = {p.name_fa: p for p in DimProvince.objects.all()}
        for row in range(PROVINCE_FIRST_ROW, PROVINCE_LAST_ROW + 1):
            name = ws.cell(row=row, column=PROVINCE_NAME_COL).value
            if not name:
                continue
            name = str(name).strip()
            province = prov_index.get(name)
            if province is None:
                province = DimProvince.objects.create(
                    code=f"prov-x-{row}", name_fa=name
                )
                prov_index[name] = province
            FactSalesProvince.objects.update_or_create(
                period=period,
                province=province,
                channel=SalesChannel.TEAM,
                defaults={
                    "sales_rial": _num(ws.cell(row=row, column=PROVINCE_SALES_COL).value),
                    "target_rial": _num(ws.cell(row=row, column=PROVINCE_TARGET_COL).value),
                },
            )
            n_prov += 1

        msg = f"Imported {n_facts} salespeople + {n_prov} province rows for {period.label} ({status})."
        self.stdout.write(self.style.SUCCESS(msg))

        if opts["approve"]:
            written = compute_period_kpis(period)
            self.stdout.write(self.style.SUCCESS(f"Computed {written} KPI rows."))
