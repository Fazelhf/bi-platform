"""
Excel export for the executive dashboards.

Builds a real .xlsx workbook (openpyxl) per section — one sheet for the KPI
table and one per detail block — so the CEO can hand the board a file instead
of a screenshot. Sheets are right-to-left and Rial columns get thousands
separators, matching how the numbers read in the app.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apps.core.models import FactKPI, KPIScope
from apps.production.models import (
    FactProduction,
    FactProductionCost,
)
from apps.sales.models import (
    ApprovalStatus,
    FactSalesMonthly,
    FactSalesProvince,
    SalesChannel,
)

# --- shared styling -------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1C1C1E")   # design-system ink
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=13)
THIN = Side(style="thin", color="E2E1DC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RIAL_FMT = "#,##0"
PCT_FMT = "0.0"

CHANNEL_LABEL = {
    SalesChannel.TEAM: "فروش همکار",
    SalesChannel.ORGANIZATIONAL: "فروش بانکی",
    SalesChannel.B2B: "فروش B2B",
}


def _sheet(wb, title, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.sheet_view.rightToLeft = True
    return ws


def _write_table(ws, headers, rows, formats=None, start_row=1):
    """Write a header row + data rows, styled, and auto-size the columns."""
    formats = formats or {}
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for r, row in enumerate(rows, start_row + 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.border = BORDER
            if c - 1 in formats:
                cell.number_format = formats[c - 1]
                cell.alignment = Alignment(horizontal="left")
    # Width from the longest cell in each column (Persian text is wide).
    for c, h in enumerate(headers, 1):
        longest = max(
            [len(str(h))] + [len(str(r[c - 1])) for r in rows if r[c - 1] is not None]
        ) if rows else len(str(h))
        ws.column_dimensions[get_column_letter(c)].width = min(42, max(12, longest + 4))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    return start_row + len(rows) + 1


def _kpi_rows(period, domain, channel=""):
    kpis = FactKPI.objects.filter(
        period=period, scope=KPIScope.COMPANY, kpi__domain=domain, channel=channel
    ).select_related("kpi").order_by("kpi__code")
    rows = []
    for k in kpis:
        rows.append([
            k.kpi.name_fa,
            k.kpi.name_en,
            float(k.actual) if k.actual is not None else None,
            float(k.target) if k.target is not None else None,
            float(k.ideal) if k.ideal is not None else None,
            float(k.deviation) if k.deviation is not None else None,
            float(k.efficiency_pct) if k.efficiency_pct is not None else None,
            k.kpi.unit,
        ])
    return rows


def _kpi_sheet(wb, period, domain, channel, title, first=False):
    ws = _sheet(wb, "شاخص‌ها", first=first)
    ws.cell(row=1, column=1, value=f"{title} — {period.label}").font = TITLE_FONT
    _write_table(
        ws,
        ["شاخص", "نام انگلیسی", "واقعی", "مطلوب", "ایده‌آل", "انحراف", "بهره‌وری (٪)", "واحد"],
        _kpi_rows(period, domain, channel),
        formats={2: RIAL_FMT, 3: RIAL_FMT, 4: RIAL_FMT, 5: RIAL_FMT, 6: PCT_FMT},
        start_row=3,
    )
    return ws


def _sales_workbook(period, channel):
    """KPIs + per-salesperson rows + provinces, for one sales channel."""
    label = CHANNEL_LABEL.get(channel, channel)
    wb = Workbook()
    _kpi_sheet(wb, period, "sales", channel, label, first=True)

    is_b2b = channel == SalesChannel.B2B
    facts = FactSalesMonthly.objects.filter(
        period=period, channel=channel, status=ApprovalStatus.APPROVED
    ).select_related("employee", "employee__team").order_by("employee__id")

    ws = _sheet(wb, "شرکت‌ها" if is_b2b else "فروشندگان")
    if is_b2b:
        headers = ["شرکت", "فروش ریالی", "مقدار (تن)", "تعداد قرارداد", "شرکت فعال",
                   "شرکت جدید", "سود", "هزینه", "تارگت", "وصول‌شده", "مانده مطالبات",
                   "فاکتورهای برنده‌شده"]
        rows = [[
            f.employee.full_name_fa, float(f.revenue_rial), float(f.quantity_ton),
            f.invoice_count, f.active_customers, f.new_customers, float(f.profit_rial),
            float(f.cost_rial), float(f.target_rial), float(f.collected_rial),
            float(f.receivables_rial), float(f.won_invoices_rial),
        ] for f in facts]
        fmts = {i: RIAL_FMT for i in [1, 6, 7, 8, 9, 10, 11]}
    else:
        headers = ["فروشنده", "تیم", "فروش ریالی", "تعداد فاکتور", "مشتری فعال",
                   "مشتری جدید", "سود", "هزینه", "تارگت", "تعداد تماس"]
        rows = [[
            f.employee.full_name_fa,
            f.employee.team.name_fa if f.employee.team else "",
            float(f.revenue_rial), f.invoice_count, f.active_customers,
            f.new_customers, float(f.profit_rial), float(f.cost_rial),
            float(f.target_rial), f.calls,
        ] for f in facts]
        fmts = {i: RIAL_FMT for i in [2, 6, 7, 8]}
        if channel == SalesChannel.TEAM:
            headers += ["پیش‌فاکتور صادره", "پیش‌فاکتور کنسل‌شده"]
            for row, f in zip(rows, facts):
                row += [float(f.proforma_issued_rial), float(f.proforma_cancelled_rial)]
            fmts[10] = RIAL_FMT
            fmts[11] = RIAL_FMT
    _write_table(ws, headers, rows, formats=fmts)

    ws = _sheet(wb, "استان‌ها")
    provinces = FactSalesProvince.objects.filter(
        period=period, channel=channel
    ).select_related("province").order_by("-sales_rial")
    _write_table(
        ws,
        ["استان", "فروش (ریال)", "تارگت (ریال)", "تحقق (٪)"],
        [[
            p.province.name_fa, float(p.sales_rial), float(p.target_rial),
            round(float(p.sales_rial) / float(p.target_rial) * 100, 1) if p.target_rial else None,
        ] for p in provinces],
        formats={1: RIAL_FMT, 2: RIAL_FMT, 3: PCT_FMT},
    )
    return wb, label


def _production_workbook(period):
    wb = Workbook()
    _kpi_sheet(wb, period, "production", "", "تولید", first=True)

    ws = _sheet(wb, "خطوط تولید")
    facts = FactProduction.objects.filter(
        period=period, status=ApprovalStatus.APPROVED
    ).select_related("machine").order_by("machine__sort_order")
    _write_table(
        ws,
        ["خط تولید", "شیفت فعال", "تولید", "ضایعات (٪)", "تعمیرات",
         "توقف خرابی", "توقف تعویض سایز", "توقف بی‌کاری"],
        [[
            f.machine.name_fa, float(f.active_shifts), float(f.output_units),
            float(f.waste_pct), float(f.repair_count),
            float(f.downtime_breakdown_shifts), float(f.downtime_sizechange_shifts),
            float(f.downtime_nowork_shifts),
        ] for f in facts],
        formats={2: RIAL_FMT, 3: PCT_FMT},
    )

    ws = _sheet(wb, "هزینه‌ها")
    costs = FactProductionCost.objects.filter(
        period=period
    ).select_related("category").order_by("category__sort_order")
    rows = [[c.category.name_fa, float(c.amount_rial)] for c in costs]
    rows.append(["جمع کل", sum(r[1] for r in rows)])
    _write_table(ws, ["دسته هزینه", "مبلغ (ریال)"], rows, formats={1: RIAL_FMT})
    return wb, "تولید"


def build_workbook(period, section: str):
    """section: team | organizational | b2b | production → (BytesIO, filename)."""
    if section == "production":
        wb, label = _production_workbook(period)
    else:
        if section not in SalesChannel.values:
            raise ValueError(f"unknown section: {section}")
        wb, label = _sales_workbook(period, section)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream, f"{label} - {period.label}.xlsx"
